import openpyxl

__all__ = ["worksheets", "workbooks"]


def worksheets(ws1, ws2):
    """
    Check if two worksheets are equivalent:
      - same name
      - same value in every cell
      - same background (fill) color in every cell
    """
    # 1) Check sheet names
    if ws1.title != ws2.title:
        return False

    # 2) Compare cell-by-cell
    max_row = max(ws1.max_row, ws2.max_row)
    max_col = max(ws1.max_column, ws2.max_column)

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell1 = ws1.cell(row=row, column=col)
            cell2 = ws2.cell(row=row, column=col)

            # Compare cell values
            if cell1.value != cell2.value:
                return False

            # Compare fill colors
            # Note: fill.fgColor can be in different formats (e.g., theme-based color),
            # but .rgb is a good quick check for many cases.
            color1 = cell1.fill.fgColor.rgb if cell1.fill.fgColor else None
            color2 = cell2.fill.fgColor.rgb if cell2.fill.fgColor else None

            if color1 != color2:
                return False

    return True


def workbooks(wb1, wb2):
    """
    Compare two workbooks to check if they are equivalent:
      - same number of sheets
      - each corresponding sheet is equivalent
    """

    # 1) Compare the number of worksheets
    if len(wb1.worksheets) != len(wb2.worksheets):
        return False

    # 2) Compare sheets one by one
    for ws1, ws2 in zip(wb1.worksheets, wb2.worksheets):
        if not worksheets(ws1, ws2):
            return False

    return True


def files(file1, file2):
    # Load workbooks
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    return workbooks(wb1, wb2)
