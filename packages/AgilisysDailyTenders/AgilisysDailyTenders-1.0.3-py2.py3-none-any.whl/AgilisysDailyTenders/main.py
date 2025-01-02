from . import Classification
from . import Bluelight
from . import ContractsFinder
from . import Proactis
import os
from datetime import datetime

directory = "Daily Tender Detail Documents"
if not os.path.exists(directory):
    os.makedirs(directory)

today_directory = os.path.join(directory, datetime.now().strftime('%Y-%m-%d_%H-%M-%S'))
if not os.path.exists(today_directory):
    os.makedirs(today_directory)

bluelight_filename = Bluelight.save(today_directory)
contractsfinder_filename = ContractsFinder.save(today_directory)
proactis_filename = Proactis.save(today_directory)

import pandas as pd
import openpyxl
from datetime import datetime
import re
import os

file_path = os.path.join(today_directory, 'Shortlisted_Tenders_' + datetime.now().strftime('%Y-%m-%d_%H-%M-%S') + '.xlsx')

def _get_link_if_exists(cell) -> str | None:
    try:
        return cell.hyperlink.target
    except AttributeError:
        return None

Bluelight_df = pd.read_excel(bluelight_filename, sheet_name='Sheet')
Bluelight_ws = openpyxl.load_workbook(bluelight_filename)['Sheet']
Bluelight_df['Extracted_URL'] = [_get_link_if_exists(Bluelight_ws.cell(row=i+2, column=3)) for i in range(len(Bluelight_df['Title']))]

ContractsFinder_df = pd.read_excel(contractsfinder_filename, sheet_name='Sheet')
ContractsFinder_ws = openpyxl.load_workbook(contractsfinder_filename)['Sheet']
ContractsFinder_df['Extracted_URL'] = [_get_link_if_exists(ContractsFinder_ws.cell(row=i+2, column=1)) for i in range(len(ContractsFinder_df['Title']))]

Proactis_df = pd.read_excel(proactis_filename, sheet_name='Sheet')
Proactis_ws = openpyxl.load_workbook(proactis_filename)['Sheet']
Proactis_df['Extracted_URL'] = [_get_link_if_exists(Proactis_ws.cell(row=i+2, column=1)) for i in range(len(Proactis_df['Title']))]

def classify_and_filter_in_place(df):
    rows_to_keep = []
    for index, row in df.iterrows():
        tender = row['Title']
        # Preprocess tender information
        preprocessed_tender = Classification.Preprocessing(tender)
        print(preprocessed_tender)

        # Apply Classification
        Bert = Classification.BertEncoding(preprocessed_tender)
        print(Bert)
        tfidf = Classification.TfIdf(preprocessed_tender)
        print(tfidf)
        patternmatching = Classification.PatternMatching(preprocessed_tender)
        print(patternmatching)
        
        # Check if row meets conditions
        if (Bert + tfidf + patternmatching) >= 2:
            rows_to_keep.append(index)  # Keep index of rows that meet conditions
        else:
            tender = row['Description']
            # Preprocess tender information
            preprocessed_tender = Classification.Preprocessing(tender)
            print(preprocessed_tender)

            # Apply Classification
            Bert = Classification.BertEncoding(preprocessed_tender)
            print(Bert)
            tfidf = Classification.TfIdf(preprocessed_tender)
            print(tfidf)
            patternmatching = Classification.PatternMatching(preprocessed_tender)
            print(patternmatching)
            
            # Check if row meets conditions
            if (Bert + tfidf + patternmatching) >= 2:
                rows_to_keep.append(index)  # Keep index of rows that meet conditions

    
    # Filter DataFrame to only keep rows that met conditions
    df = df.loc[rows_to_keep]
    return df

Bluelight_df = classify_and_filter_in_place(Bluelight_df)
ContractsFinder_df = classify_and_filter_in_place(ContractsFinder_df)
Proactis_df = classify_and_filter_in_place(Proactis_df)

with pd.ExcelWriter(file_path) as writer:
    Bluelight_df.to_excel(writer, sheet_name='Bluelight', index=False)
    ContractsFinder_df.to_excel(writer, sheet_name='ContractsFinder', index=False)
    Proactis_df.to_excel(writer, sheet_name='Proactis', index=False)