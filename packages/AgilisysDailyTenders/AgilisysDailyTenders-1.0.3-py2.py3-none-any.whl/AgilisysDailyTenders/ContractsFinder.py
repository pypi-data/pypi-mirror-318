import asyncio
from playwright.async_api import Playwright, async_playwright
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import openpyxl
import os
from openpyxl.styles import Font
from openpyxl import Workbook

def save (todays_directory):
    contractsfinder_filename = os.path.join(todays_directory,f"contractsfinder_extracted_data_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")

    async def run(playwright: Playwright) -> None:
        # Launch the browser
        browser = await playwright.chromium.launch(headless=True)
        context = await browser.new_context()
        page = await context.new_page()

        # Navigate to the Contracts Finder website
        await page.goto("https://www.contractsfinder.service.gov.uk/Search/Results")

        # Accept cookies and hide messages
        await page.get_by_role("button", name="Accept analytics cookies").click()
        await page.get_by_role("button", name="Hide this message").click()

        # Date handling
        today = datetime.now()
        to_date = today
        if to_date.weekday() == 0:  # If today is Monday
            from_date = to_date - timedelta(days=3)  # Set from_date to Friday
        else:
            from_date = to_date - timedelta(days=1)  # Set from_date to yesterday

        to_date_str = to_date.strftime("%d")
        from_date_str = from_date.strftime("%d")
        to_month_str = to_date.strftime("%m")
        from_month_str = from_date.strftime("%m")
        to_year_str = to_date.strftime("%Y")
        from_year_str = from_date.strftime("%Y")

        # Fill in the CPV codes and date filters
        await page.get_by_role("button", name="Industry CPV code").click()
        await page.get_by_role("button", name="Browse CPV codes").click()
        await page.get_by_text("32000000: Radio, television,").click()
        await page.get_by_text("48000000: Software package").click()
        await page.get_by_text("50000000: Repair and").click()
        await page.get_by_label("64000000: Postal and").check()
        await page.get_by_label("66000000: Financial and").check()
        await page.get_by_text("72000000: IT services:").click()
        await page.get_by_label("73000000: Research and").check()
        await page.get_by_text("79000000: Business services:").click()
        await page.get_by_text("98000000: Other community,").click()
        await page.get_by_role("button", name="Add CPV code(s) to filter").click()

        await page.get_by_role("button", name="Date range").click()
        await page.fill("#published_date_from-day", from_date_str)
        await page.fill("#published_date_from-month", from_month_str)
        await page.fill("#published_date_from-year", from_year_str)

        await page.wait_for_selector("#published_date_to-day")
        await page.fill("#published_date_to-day", to_date_str)
        await page.fill("#published_date_to-month", to_month_str)
        await page.fill("#published_date_to-year", to_year_str)

        # Click the update results button
        await page.get_by_role("button", name="Update results").click()

        # Wait for results to load
        await page.wait_for_load_state("networkidle")

        workbook = Workbook()
        sheet = workbook.active
        # Write headers
        sheet.append([ 
            "Title", "Company Name", "Description", "Procurement Stage", 
            "Notice Status", "Closing", "Contract Location", 
            "Contract Value", "Publication Date", "Category key"
        ])

        # Function to scrape data from the current page
        async def scrape_data():
            soup = BeautifulSoup(await page.content(), "html.parser")
            search_results = soup.find_all("div", class_="search-result")

            # Write data to Excel
            next_row = sheet.max_row + 1
            for search_result in search_results:
                # Extract title and link
                title_element = search_result.find("h2")
                title = title_element.get_text(strip=True)
                link = title_element.find("a")["href"]

                # Extract company name
                company_name_element = search_result.find("div", class_="search-result-sub-header")
                company_name = company_name_element.get_text(strip=True) if company_name_element else "N/A"

                # Extract additional details
                procurement_stage, notice_status, closing, contract_location, contract_value, publication_date = ("",) * 6
                for entry in search_result.find_all("div", class_="search-result-entry"):
                    label_tag = entry.find("strong")
                    label = label_tag.get_text(strip=True)
                    value = entry.get_text(strip=True).replace(label, "", 1).strip()
                    if label == "Procurement stage":
                        procurement_stage = value
                    elif label == "Notice status":
                        notice_status = value
                    elif label == "Closing":
                        closing = value
                    elif label == "Contract location":
                        contract_location = value
                    elif label == "Contract value":
                        contract_value = value
                    elif label == "Publication date":
                        publication_date = value

                # Click the title link to open the contract page
                await page.goto(link)
                await page.wait_for_load_state("networkidle")

                # Extract Industry and Description from the contract page
                contract_page_soup = BeautifulSoup(await page.content(), "html.parser")

                # Extract industry information
                industry = contract_page_soup.find("h4", string="Industry")
                industry_list = []
                if industry:
                    ul = industry.find_next("ul")
                    for li in ul.find_all("li"):
                        industry_list.append(li.get_text(strip=True))
                industry_text = ", ".join(industry_list) if industry_list else "N/A"

                # Extract description specifically from the "Description" block
                description_text = ""

                # Look for any <h3> with the text "Description"
                description_section = contract_page_soup.find("h3", string="Description")

                if description_section:
                    temp = description_section.find_next("p")
                    p_tags = temp.find_next("p")
                    description_text = p_tags.get_text(" ", strip=True) if p_tags else "No description available"

                # Write to Excel
                sheet.cell(row=next_row, column=1).value = title
                sheet.cell(row=next_row, column=1).hyperlink = link
                sheet.cell(row=next_row, column=1).font = Font(color="0000EE", underline="single")
                sheet.cell(row=next_row, column=2).value = company_name
                sheet.cell(row=next_row, column=3).value = description_text
                sheet.cell(row=next_row, column=4).value = procurement_stage
                sheet.cell(row=next_row, column=5).value = notice_status
                sheet.cell(row=next_row, column=6).value = closing
                sheet.cell(row=next_row, column=7).value = contract_location
                sheet.cell(row=next_row, column=8).value = contract_value
                sheet.cell(row=next_row, column=9).value = publication_date
                sheet.cell(row=next_row, column=10).value = industry_text
                next_row += 1

        # Scrape the first page
        await scrape_data()

        # Pagination
        while True:
            # Check if there's a next page
            next_page_button = await page.query_selector("a.standard-paginate-next")
            if next_page_button:
                await next_page_button.click()
                await page.wait_for_load_state("networkidle")  # Wait for the next page to load
                await scrape_data()  # Scrape the next page
            else:
                break  # No more pages

        workbook.save(contractsfinder_filename)
        print(f"Data successfully written to {contractsfinder_filename}.")

        await context.close()
        await browser.close()

    async def main() -> None:
        async with async_playwright() as playwright:
            await run(playwright)

    asyncio.run(main())
    return contractsfinder_filename