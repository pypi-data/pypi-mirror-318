import asyncio
from playwright.async_api import Playwright, async_playwright
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook
import os

def save(today_directory):
    proactis_filename = os.path.join(today_directory, f"proactis_extracted_data_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")

    async def run(playwright: Playwright) -> None:
        # Launch the browser
        browser = await playwright.chromium.launch(headless=True)
        context = await browser.new_context()
        page = await context.new_page()
        
        # Navigate to the website
        await page.goto("https://procontract.due-north.com/Opportunities/Index?resetFilter=True&applyFilter=True&p=2241eb95-058a-e511-80f7-000c29c9ba21&v=1")

        # Wait for the cookie consent dialog and click "Accept All" if it appears
        try:
            await page.get_by_role("button", name="Accept all").click()
            # await page.wait_for_selector("text=Accept all", timeout=10000)
            # await page.evaluate("document.querySelector('div.js-cookie-consent-dialog').scrollIntoView();")
            # await page.click("text=Accept all", force=True)
        except Exception as e:
            print("Cookie consent dialog not found or could not be clicked:", e)
            # await page.evaluate("document.querySelector('.js-cookie-consent-dialog').style.display='none'")

        today = datetime.now()
        to_date = today
        from_date = to_date - timedelta(days=3 if to_date.weekday() == 0 else 1)

        to_date_str = to_date.strftime("%d/%m/%Y")
        from_date_str = from_date.strftime("%d/%m/%Y")

        # print(f"To Date: {to_date_str}, From Date: {from_date_str}")

        # await page.fill("#StartEndDateFilter_1__StartDate", from_date_str)
        # await page.fill("#StartEndDateFilter_1__EndDate", to_date_str)
        
        # await page.get_by_text("Update", exact=True).click()
        # await page.wait_for_load_state("networkidle")

        while True:

            await page.fill("#StartEndDateFilter_1__StartDate", from_date_str)
            await page.fill("#StartEndDateFilter_1__EndDate", to_date_str)
            
            await page.get_by_text("Update", exact=True).click()
            await page.wait_for_load_state("networkidle")

            check_from_date_str = await page.get_attribute("#StartEndDateFilter_1__StartDate", "value")
            check_to_date_str = await page.get_attribute("#StartEndDateFilter_1__EndDate", "value")

            if(check_from_date_str == from_date_str and check_to_date_str == to_date_str):
                break

        workbook = Workbook()
        sheet = workbook.active
        headers = ["Title", "Buyer", "Expression Start", "Expression End", "Estimated Value", "Category Key", "Description"]  # Add headers based on your table structure
        sheet.append(headers)

        # Function to scrape data from the current page
        async def scrape_data():
            soup = BeautifulSoup(await page.content(), "html.parser")
            table = soup.find('table')
            rows = []

            for row in table.find_all('tr')[1:]:
                cells = row.find_all(['td', 'th'])
                row_data = []
                for cell in cells:
                    anchor = cell.find('a')
                    if anchor:
                        link = anchor['href']
                        if not link.startswith('http'):
                            link = "https://procontract.due-north.com/" + link
                        row_data.append((anchor.get_text().strip(), link))
                    else:
                        row_data.append(cell.get_text().strip())
                rows.append(row_data)

            if "There is no data available." in row_data[0]:
                # print("Table indicates no data is available.")
                return

            # Write data to Excel
            next_row = sheet.max_row + 1
            for row_data in rows:
                title, link = row_data[0]  # Extract title and link
                buyer = row_data[1]
                expression_start = row_data[2]
                expression_end = row_data[3]
                estimated_value = row_data[4]

                # Write title as a hyperlink in the first column
                sheet.cell(row=next_row, column=1).hyperlink = link
                sheet.cell(row=next_row, column=1).value = title
                sheet.cell(row=next_row, column=1).font = Font(color="0000EE", underline="single")
                
                # Write other details (Buyer, Expression Start, etc.)
                sheet.cell(row=next_row, column=2).value = buyer
                sheet.cell(row=next_row, column=3).value = expression_start
                sheet.cell(row=next_row, column=4).value = expression_end
                sheet.cell(row=next_row, column=5).value = estimated_value

                # Click on the link to open the tender details page
                await page.goto(link)
                await page.wait_for_load_state("networkidle")

                # Extract category and description from the detailed page
                category = await page.locator('div.row:has-text("Categories") .cell400').inner_text()
                description = await page.locator('div.row:has-text("Description") .cell400').inner_text()

                # Write category key and description to Excel
                sheet.cell(row=next_row, column=6).value = category
                sheet.cell(row=next_row, column=7).value = description

                next_row += 1
                
                # Go back to the main page
                await page.go_back()

        # Scrape the first page
        await scrape_data()

        # Pagination
        while True:
            # Check for the next page button
            next_page_button = await page.query_selector("a.pager.next")
            if next_page_button:
                await next_page_button.click()
                await page.wait_for_load_state("networkidle")  # Wait for the next page to load
                await scrape_data()  # Scrape the next page
            else:
                break  # No more pages

        workbook.save(proactis_filename)
        print(f"Data successfully written to {proactis_filename}.")

        await context.close()
        await browser.close()

    async def main() -> None:
        async with async_playwright() as playwright:
            await run(playwright)

    asyncio.run(main())
    return proactis_filename