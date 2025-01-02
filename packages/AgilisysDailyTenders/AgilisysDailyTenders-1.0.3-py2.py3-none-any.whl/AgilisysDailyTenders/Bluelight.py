import asyncio
from playwright.async_api import Playwright, async_playwright
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook
import os

def save(today_directory):
    filename = f"Bluelight_extracted_data_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    bluelight_filename = os.path.join(today_directory,filename)
    
    async def run(playwright: Playwright) -> None:
        # Playwright Part
        browser = await playwright.chromium.launch(headless=True)
        context = await browser.new_context()
        page = await context.new_page()
        await page.goto("https://uk.eu-supply.com/ctm/supplier/publictenders?B=BLUELIGHT")

        today = datetime.now()
        to_date = today

        if to_date.weekday() == 0:
            from_date = to_date - timedelta(days=3)
        else:
            from_date = to_date - timedelta(days=1)

        to_date_str = to_date.strftime("%d/%m/%Y")
        from_date_str = from_date.strftime("%d/%m/%Y")

        print(f"From Date: {from_date_str}, To Date: {to_date_str}")

        await page.fill("#SearchFilter_FromDate", from_date_str)
        await page.fill("#SearchFilter_ToDate", to_date_str)

        await page.get_by_text("Search", exact=True).click()
        await page.wait_for_load_state("networkidle")
        
        workbook = Workbook()
        sheet = workbook.active
        headers = ["Quote/Tender Id", "Reference", "Title", "Description", "Date of Publication", "Response Deadline", 
                   "Process", "Buyers", "Countries", "Category Key"]
        sheet.append(headers)

        # Function to scrape data from the current page
        async def scrape_data():
            soup = BeautifulSoup(await page.content(), "html.parser")
            table = soup.find('table')
            # rows = []

            for row in table.find_all('tr')[1:]:  # Skip the header row
                cells = row.find_all('td')
                row_data = []
                
                # Check each column, extract data or set "N/A" if missing
                for cell in cells:
                    # Extract basic tender information
                    tender_detail = cell.get_text().strip() if cell.get_text().strip() else "N/A"
                    row_data.append(tender_detail)

                # Check if there's an anchor tag in the second column (Name column with link)
                tender_link_tag = cells[2].find('a')  # Updated: now it's the Name column (index 2)
                if tender_link_tag and 'href' in tender_link_tag.attrs:
                    tender_link = tender_link_tag['href']
                    if not tender_link.startswith('http'):
                        tender_link = "https://uk.eu-supply.com" + tender_link
                else:
                    tender_link = None

                # Add additional details like Description and Category Key by visiting each link
                if tender_link:
                    row_data.append(tender_link)
                    print(f"Opening link: {tender_link}")  # Debugging line
                    await page.goto(tender_link)
                    await page.wait_for_load_state("networkidle")

                    # Extract Description (using the class "ctm-multi-line")
                    description_elem = await page.query_selector('div.ctm-content-label:has-text("Detailed description:") + p.ctm-multi-line')
                    if description_elem:
                        description = await description_elem.text_content()
                    else:
                        description = "No description available"
                    print(f"Description: {description}")  # Debugging line

                    # Extract Category Key (using the class "text-underline")
                    category_elem = await page.query_selector('div.ctm-content-label:has-text("CPV codes:") + p span.text-underline')
                    if category_elem:
                        category_key = await category_elem.text_content()
                    else:
                        category_key = "No category key available"
                    print(f"Category Key: {category_key}")  # Debugging line

                    # Write data to Excel
                    next_row = sheet.max_row + 1
                    sheet.cell(row=next_row, column=1).value = row_data[0]  # Quote/Tender Id
                    sheet.cell(row=next_row, column=2).value = row_data[1]  # Reference
                    sheet.cell(row=next_row, column=3).value = row_data[2]  # Name
                    sheet.cell(row=next_row, column=3).hyperlink = row_data[8]
                    sheet.cell(row=next_row, column=3).font = Font(color="0000EE", underline="single")
                    sheet.cell(row=next_row, column=4).value = description  # Description
                    sheet.cell(row=next_row, column=5).value = row_data[3]  # Publication Date
                    sheet.cell(row=next_row, column=6).value = row_data[4]  # Response Deadline
                    sheet.cell(row=next_row, column=7).value = row_data[5]  # Process
                    sheet.cell(row=next_row, column=8).value = row_data[6]  # Buyers
                    sheet.cell(row=next_row, column=9).value = row_data[7]  # Countries
                    sheet.cell(row=next_row, column=10).value = category_key  # Category Key

                    # print(f"Written tender {row_data[1]} to Excel")

                    # Go back to the main page after processing each tender
                    await page.go_back()
                   

        # Scrape the first page
        await scrape_data()

        # Pagination
        while True:
            # Check for the next page button
            next_page_button = await page.query_selector("a.pager-action:not(.state-active)")
            if next_page_button:
                await next_page_button.click()
                await page.wait_for_load_state("networkidle")  # Wait for the next page to load
                await scrape_data()  # Scrape the next page
            else:
                break  # No more pages

        # Save the workbook
        workbook.save(bluelight_filename)
        print(f"Data successfully written to {bluelight_filename}.")

        await context.close()
        await browser.close()

    async def main() -> None:
        async with async_playwright() as playwright:
            await run(playwright)

    asyncio.run(main())
    return bluelight_filename