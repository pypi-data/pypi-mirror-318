"""
xlsxtools: Utility functions for writing and manipulating Excel files

- write_dataframe_to_excel() : write a Pandas DataFrame to an Excel file, optionally calling
    auto_size_excel_column_widths() and filter_and_freeze_excel()
- auto_size_excel_column_widths() : Reformat an Excel file, adjusting the column widths to fit their contents
- filter_and_freeze_excel() : Clean up an Excel file by setting a filter on the first row,
    splitting & freezing the panes


Releases:
0.1.0, 2021-10-26
    * Initial release

0.2.0, 2025-01-02
    * Move from pandas.ExcelWriter() to DataFrame.to_excel() (transparent to user)
    * Add optional `tabname` argument to auto_size_excel_column_widths() and filter_and_freeze_excel()
"""

from __future__ import annotations

import pathlib
from pathlib import Path
import pandas as pd
import openpyxl

__version__ = '0.2.0'


def write_dataframe_to_excel(df: pd.DataFrame, filepath: str | Path, tabname: str = 'data', index: bool = True,
                             auto_size_column_widths: bool = False, filter_and_freeze: str | None = None):
    df.to_excel(excel_writer=filepath, sheet_name=tabname, index=index)

    if auto_size_column_widths or filter_and_freeze:
        wb = openpyxl.load_workbook(filepath)

        if auto_size_column_widths:
            auto_size_excel_column_widths(wb)
        if filter_and_freeze:
            filter_and_freeze_excel(wb, freeze_location=filter_and_freeze)

        wb.save(filepath)


def auto_size_excel_column_widths(workbook: str | Path | openpyxl.workbook.workbook.Workbook,
                                  tabname: str | None = None):
    # 'workbook' must be a path pointing to the Excel file (str or pathlib.Path)
    # or an openpyxl.workbook.workbook.Workbook
    if isinstance(workbook, str):
        xlsxpath = workbook
        workbook = openpyxl.load_workbook(workbook)
        savefile = True
    elif isinstance(workbook, (Path, pathlib.WindowsPath, pathlib.PosixPath)):
        # openpyxl.load_workbook takes a str as a path, not a Path object
        xlsxpath = str(workbook)

        workbook = openpyxl.load_workbook(str(workbook))
        savefile = True
    elif isinstance(workbook, openpyxl.workbook.workbook.Workbook):
        savefile = False
    else:
        raise TypeError(f'workbook object is not a valid type: {type(workbook)}')

    if tabname is None:
        worksheet = workbook.active
    else:
        worksheet = workbook[tabname]

    dims = {}
    for rownum, row in enumerate(worksheet.rows):
        if rownum == 0:
            padding = 1
        else:
            padding = 1

        for cell in row:
            if cell.value:
                if cell.font.b:
                    scaling = 1.1
                else:
                    scaling = 1
                try:
                    dims[cell.column_letter] = max(dims.get(cell.column_letter, 0),
                                                   int(round(len(cell.value)*scaling)))
                except TypeError:
                    dims[cell.column_letter] = max(dims.get(cell.column_letter, 0),
                                                   int(round(len(str(cell.value))*scaling)))

    for column_letter, width in dims.items():
        worksheet.column_dimensions[column_letter].width = max(width + padding, 9)

    # if a file path was passed in (not a Workbook object), we save the manipulated file
    if savefile:
        workbook.save(filename=xlsxpath)

    return dims


def filter_and_freeze_excel(workbook, freeze_location='A2',
                            tabname: str | None = None):
    """
    Set filter by first row, freeze panes
    """
    # 'workbook' must be a path pointing to the Excel file (str or pathlib.Path)
    # or an openpyxl.workbook.workbook.Workbook
    savefile = False
    if isinstance(workbook, str):
        xlsxpath = workbook
        workbook = openpyxl.load_workbook(workbook)
        savefile = True
    elif isinstance(workbook, (Path, pathlib.WindowsPath, pathlib.PosixPath)):
        # openpyxl.load_workbook takes a str as a path, not a Path object
        xlsxpath = str(workbook)

        workbook = openpyxl.load_workbook(str(workbook))
        savefile = True
    elif not isinstance(workbook, openpyxl.workbook.workbook.Workbook):
        raise TypeError(f'workbook object is not a valid type: {type(workbook)}')

    if tabname is None:
        worksheet = workbook.active
    else:
        worksheet = workbook[tabname]

    c = worksheet[freeze_location]
    worksheet.freeze_panes = c

    row_count = worksheet.max_row
    column_count = worksheet.max_column

    worksheet.auto_filter.ref = "A1:%s%i" % (openpyxl.utils.get_column_letter(column_count), row_count)

    # if a file path was passed in (not a Workbook object), we save the manipulated file
    if savefile:
        workbook.save(filename=xlsxpath)
