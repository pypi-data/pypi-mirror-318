#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/4/19 19:01
# @Author  : yaitza
# @Email   : yaitza@foxmail.com
import json
from openpyxl import *


class XlsxHandler():
	def __init__(self, file_path):
		self.file_path = file_path

	def load(self):
		raw = []
		workbook = load_workbook(self.file_path)
		table = workbook.active
		for row in range(2, table.max_row + 1):
			tester = {"name": None, "request": {}, "validate": [], "expect": None}
			for col in range(table.max_column):
				col_title = table.cell(1, col + 1).value
				if col_title in ['url', 'method', 'params']:
					request = tester.get("request")
					if col_title == "params":
						request_params = eval(table.cell(row, col + 1).value)
						request.update({table.cell(1, col + 1).value: request_params})
					else:
						request.update({table.cell(1, col + 1).value: table.cell(row, col + 1).value})
					tester.update({"request": request})
				elif col_title == "validate":
					if table.cell(row, col + 1).value is not None and table.cell(row, col + 1).value != '':
						tester.get("validate").append(json.loads(table.cell(row, col + 1).value))
				elif col_title == "expect":
					tester.update({table.cell(1, col + 1).value: json.loads(table.cell(row, col + 1).value)})
				else:
					tester.update({table.cell(1, col + 1).value: table.cell(row, col + 1).value})
			if not ("status" in tester.keys() and tester.get("status") == 0) :
				raw.append({"test": tester})
		return raw
